from odoo import fields, models, api, _
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import base64
from io import BytesIO


class inventory_stock_valuation_wizard_spt(models.TransientModel):
    _name = "inventory.stock.valuation.wizard.spt"

    warehouses_ids = fields.Many2many(
        'stock.warehouse', 'warehouses_comapny_rel', 'warehouses_id', 'partner_comapny_id', string='Warehouse')
    companys_id = fields.Many2one('res.company', string='Company')
    location_id = fields.Many2one('stock.location', string='Location')

    filter_by = fields.Selection(
        [('product', 'Product'), ('category', 'Category')], string='Filter By')

    start_period = fields.Date(string='Start Period')
    end_period = fields.Date(string='End Period')
    summary = fields.Boolean(string='Summary')

    category_ids = fields.Many2many(
        'product.category', 'product_category_rel', 'category_id', 'filter_id', string='Category')

    product_ids = fields.Many2many(
        'product.product', 'product_product_rel', 'product_id', 'filter_id', string='Product')

    file = fields.Binary('File')

    def action_inventory_stock_valuation_pdf_report(self):
        return self.env.ref('Inventory_stock_valuation.report_action_inventory_stock_valuation_pdf_report').report_action(self)

    # def get_report_data(self):

    def action_inventory_stock_valuation_excel_report(self):
        for rec in self:
            active_id = self.id
            f_name = 'Inventory Stock Valuation Report'
            workbook = Workbook()
            alignment = Alignment(horizontal='center',vertical='center',text_rotation=0)
            sheet = workbook.create_sheet(title=rec.companys_id.name ,index=0)
            sheet.merge_cells('D1:G2')
            sheet.cell(row=1, column=4).value = 'Inventory Stock Valuation'

            sheet.cell(row=4, column=1).value = 'Start Date'
            sheet.cell(row=4, column=2).value = 'End Date'
            sheet.cell(row=4, column=3).value = 'Comapny'
            sheet.cell(row=4, column=4).value = 'Warehoues'
            sheet.cell(row=4, column=5).value = 'Currency'

            for inventory_stock in self:
                sheet.cell(row=5, column=1).value = inventory_stock.start_period
                sheet.cell(row=5, column=2).value = inventory_stock.end_period
                sheet.cell(row=5, column=3).value = inventory_stock.companys_id.name
                sheet.cell(row=5, column=5).value = inventory_stock.companys_id.currency_id.name
                warehouse_list = []
                for warehouses_id in inventory_stock.warehouses_ids:
                    warehouse_list.append(warehouses_id.name)
                sheet.cell(row= 5, column=4).value = ','.join(warehouse_list)

                sheet.column_dimensions['A'].width = 15
                sheet.column_dimensions['B'].width = 25
                sheet.column_dimensions['C'].width = 25
                sheet.column_dimensions['D'].width = 25
                sheet.column_dimensions['E'].width = 10

                sheet.column_dimensions['A'].alignment = alignment
                sheet.column_dimensions['B'].alignment = alignment
                sheet.column_dimensions['C'].alignment = alignment
                sheet.column_dimensions['D'].alignment = alignment
                sheet.column_dimensions['E'].alignment = alignment    
                 
            
                sheet.cell(row=7, column=1).value = 'Default Code'
                sheet.cell(row=7, column=2).value = 'Name'
                sheet.cell(row=7, column=3).value = 'Category'
                row_index = 8

                for product in inventory_stock.product_ids:
                    sheet.cell(row=row_index, column=1).value = product.default_code
                    sheet.cell(row=row_index, column=2).value = product.name
                    sheet.cell(row=row_index, column=3).value = product.categ_id.name
                    row_index += 1
            

        fp = BytesIO()
        workbook.save(fp)
        fp.seek(0)
        data = fp.read()
        fp.close()
        rec.file = base64.b64encode(data)

        return {
            'type': 'ir.actions.act_url',
            'url':   'web/content/?model=inventory.stock.valuation.wizard.spt&download=true&field=file&id=%s&filename=%s.xlsx' % (active_id ,f_name),
            'target': 'self',
        }
